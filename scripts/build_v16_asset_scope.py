#!/usr/bin/env python3
"""Freeze the official 129-meter service scope without copying raw AMI."""
from openpyxl import load_workbook

from v16_common import DATA, OFFICIAL, canonical_sha, require, sha256_file, write_json


def main() -> None:
    source = next(OFFICIAL.glob("1-1_*.xlsx"))
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    fields = ("line", "meter_id", "section", "multiplier", "supply", "contract_kw", "tariff", "use", "der", "product", "industry")
    assets = []
    for values in sheet.iter_rows(min_row=3, max_col=11, values_only=True):
        if not values[0]:
            continue
        row = {key: ("" if value is None else value) for key, value in zip(fields, values)}
        row["streetlight_eligible"] = "가로등" in str(row["tariff"]) and "가로등" in str(row["use"])
        row["expected_phase_count"] = 1 if "단상" in str(row["supply"]) else 3
        assets.append(row)
    eligible = [row for row in assets if row["streetlight_eligible"]]
    require(len(assets) == 129, "BLOCKED_OFFICIAL_ASSET_COUNT_DRIFT")
    require(len(eligible) == 5, "BLOCKED_STREETLIGHT_SCOPE_DRIFT")
    require({row["meter_id"] for row in eligible} == {"B-L-9", "B-L-12", "B-L-13", "B-L-14", "B-L-35"}, "BLOCKED_STREETLIGHT_ID_DRIFT")
    payload = {
        "schema_version": "lightguard.v16.asset-scope.1",
        "status": "PRE_OUTCOME_FROZEN",
        "source_file_sha256": sha256_file(source),
        "official_asset_count": len(assets),
        "streetlight_eligible_count": len(eligible),
        "out_of_scope_count": len(assets) - len(eligible),
        "eligibility_rule": "tariff contains 가로등 AND use contains 가로등",
        "eligible_assets": eligible,
        "all_asset_scope_sha256": canonical_sha(assets),
        "raw_ami_written": False,
    }
    payload["registry_sha256"] = canonical_sha(payload)
    write_json(DATA / "v16_asset_scope_registry.json", payload)


if __name__ == "__main__":
    main()
