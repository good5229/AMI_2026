#!/usr/bin/env python3
"""Extract six traceable event windows from the ignored B-line MDMS workbook."""

from __future__ import annotations

import csv
import hashlib
import json
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from context_common import ROOT, utc_now, write_json


def parse_time(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    try:
        return datetime.fromisoformat(str(value).strip())
    except ValueError:
        return None


def scalar(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def main() -> None:
    source = next((ROOT / "official_docs" / "AMI Data Sample").glob("*B*xlsx"), None)
    if source is None:
        raise FileNotFoundError("Ignored B-line MDMS workbook is unavailable")
    events_path = ROOT / "lightguard_app" / "assets" / "data" / "ami_events.csv"
    with events_path.open(encoding="utf-8-sig", newline="") as handle:
        events = list(csv.DictReader(handle))
    if len(events) != 6:
        raise RuntimeError(f"Expected six competition AMI events, found {len(events)}")

    targets = []
    by_meter = defaultdict(list)
    for event in events:
        start = datetime.fromisoformat(event["first_sample"])
        end = datetime.fromisoformat(event["last_sample"])
        target = {
            "meter_id": event["meter_id"],
            "date": start.date().isoformat(),
            "event_start": start,
            "event_end": end,
            "window_start": start - timedelta(hours=2),
            "window_end": end + timedelta(hours=2),
            "rows": [],
        }
        targets.append(target)
        by_meter[target["meter_id"]].append(target)

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    for source_row, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        meter = str(row[0]).strip() if row[0] is not None else ""
        if meter not in by_meter:
            continue
        timestamp = parse_time(row[1])
        if timestamp is None:
            continue
        for target in by_meter[meter]:
            if target["window_start"] <= timestamp <= target["window_end"]:
                target["rows"].append({
                    "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                    "meter_id": meter,
                    "i1": scalar(row[13]),
                    "i2": scalar(row[14]),
                    "i3": scalar(row[15]),
                    "active_energy_kwh": scalar(row[3]),
                    "source_row": source_row,
                })

    output_dir = ROOT / "lightguard_app" / "assets" / "data" / "ami_event_windows"
    output_dir.mkdir(parents=True, exist_ok=True)
    manifest_events = []
    fieldnames = ["timestamp", "meter_id", "i1", "i2", "i3", "active_energy_kwh", "source_row"]
    for target in targets:
        filename = f"{target['meter_id']}_{target['date']}.csv"
        path = output_dir / filename
        if not target["rows"]:
            raise RuntimeError(f"No source rows found for {filename}")
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(target["rows"])
        manifest_events.append({
            "file": filename,
            "meter_id": target["meter_id"],
            "date": target["date"],
            "event_start": target["event_start"].isoformat(sep=" "),
            "event_end": target["event_end"].isoformat(sep=" "),
            "window_start": target["window_start"].isoformat(sep=" "),
            "window_end": target["window_end"].isoformat(sep=" "),
            "row_count": len(target["rows"]),
            "source_row_min": min(row["source_row"] for row in target["rows"]),
            "source_row_max": max(row["source_row"] for row in target["rows"]),
        })

    source_sha = hashlib.sha256(source.read_bytes()).hexdigest()
    write_json(output_dir / "replay_manifest.json", {
        "schema_version": "lightguard-ami-replay-v0.3",
        "source_kind": "anonymized_competition_ami",
        "source_workbook_sha256": source_sha,
        "source_sheet": sheet.title,
        "extracted_at": utc_now(),
        "fabrication_policy": "source rows only; no interpolation; blanks preserved",
        "events": manifest_events,
    })
    print(f"AMI replay windows: {len(manifest_events)} files, {sum(x['row_count'] for x in manifest_events)} source rows")


if __name__ == "__main__":
    main()
